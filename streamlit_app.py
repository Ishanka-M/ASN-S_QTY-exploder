"""
ASN S_QTY Exploder
------------------
S_QTY column එකේ අගයෙන් line ගාන හදනවා. S_QTY = 10 නම් line 10ක්,
හැම line එකකම S_QTY = 1. අනිත් හැම column එකකම details එහෙමම තියෙනවා.

Outputs:
  1) Exploded Excel (.xlsx)
  2) Summary PDF  — CLIENT_CODE + QR, DISPLAY_ASN_NUMBER + QR,
     සහ item-wise totals table

HU_ID generate: Keep original / None / Letters+Number / DISPLAY_ITEM_NUMBER+Number
Sheet name එක මොකක් වුණත් S_QTY column එකෙන් sheet එක auto-detect වෙනවා.
"""

import io
import copy
from functools import lru_cache
from collections import OrderedDict

import streamlit as st
from openpyxl import load_workbook
import qrcode
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table,
                                TableStyle, Image, PageBreak)
from reportlab.graphics.barcode import createBarcodeDrawing

st.set_page_config(page_title="ASN S_QTY Exploder", page_icon="📦", layout="wide")

SUMMARY_FIELDS = ["CLIENT_CODE", "DISPLAY_ASN_NUMBER", "DISPLAY_ITEM_NUMBER",
                  "LOT_NUMBER", "QUANTITY", "UOM", "S_QTY", "S_UOM",
                  "PACKAGE_TYPE", "GROSS_WEIGHT", "NET_WEIGHT"]

# ---------- helpers ----------

def find_col(ws, name):
    target = name.strip().upper()
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is not None and str(v).strip().upper() == target:
            return c
    return None


def is_blank_row(ws, r):
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=r, column=c).value not in (None, ""):
            return False
    return True


def to_int_qty(val):
    if val in (None, ""):
        return 1
    try:
        n = int(float(val))
    except (ValueError, TypeError):
        return 1
    return n if n >= 1 else 1


def to_num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def build_hu(prefix, psep, number, sep, counter, pad):
    """HU_ID = {prefix}{psep}{number}{sep}{counter}. e.g. ACIS-CHEMICAL + - + 260613 + . + 1 -> ACIS-CHEMICAL-260613.1"""
    cnt = str(counter).zfill(pad) if pad and pad > 0 else str(counter)
    if number in (None, ""):
        return f"{prefix}{sep}{cnt}"
    return f"{prefix}{psep}{number}{sep}{cnt}"


def explode_sheet(ws, sqty_col, line_col=None, renumber_line=True,
                  hu_col=None, item_col=None, hu_cfg=None):
    """
    Fast explode: iter_rows වලින් කියවලා, data rows delete කරලා, append වලින් ආයෙ ලියනවා.
    Header row 1 (styling එක්ක) රැකෙනවා. 'General' නොවන number formats ආයෙ apply වෙනවා.
    """
    max_col = ws.max_column

    # 'General' නොවන number formats විතරක් මතක තියාගන්නවා (reapply සඳහා)
    fmts = {}
    for c in range(1, max_col + 1):
        nf = ws.cell(row=2, column=c).number_format
        if nf and nf != "General":
            fmts[c] = nf

    # data rows fast read
    src_rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True)
                if any(v not in (None, "") for v in r)]

    # explode
    exploded = []
    for values in src_rows:
        n = to_int_qty(values[sqty_col - 1])
        base = list(values)
        base[sqty_col - 1] = 1
        for _ in range(n):
            exploded.append(list(base))

    if renumber_line and line_col is not None:
        for i, nv in enumerate(exploded, start=1):
            nv[line_col - 1] = i

    if hu_col is not None and hu_cfg and hu_cfg.get("mode") != "keep":
        mode = hu_cfg["mode"]
        start = int(hu_cfg.get("start", 1))
        number = hu_cfg.get("number", "")
        psep = hu_cfg.get("psep", "")
        sep = hu_cfg.get("sep", "")
        pad = hu_cfg.get("pad", 0)
        counters = {}  # prefix එකකට වෙන වෙනම counter — item මාරු වුණාම 1 ඉඳන් ආයෙ
        for nv in exploded:
            if mode == "none":
                nv[hu_col - 1] = None
            else:
                if mode == "item":
                    pv = nv[item_col - 1] if item_col else None
                    prefix = "" if pv is None else str(pv)
                else:
                    prefix = hu_cfg.get("letters", "")
                key = (prefix, number)
                c = counters.get(key, start)
                nv[hu_col - 1] = build_hu(prefix, psep, number, sep, c, pad)
                counters[key] = c + 1

    # පරණ data rows එකපාරටම මකනවා (header row 1 රැකෙනවා)
    if ws.max_row >= 2:
        ws.delete_rows(2, ws.max_row - 1)

    # fast write
    for nv in exploded:
        ws.append(nv)

    # number formats reapply (few cols විතරයි → fast)
    if fmts:
        for row in ws.iter_rows(min_row=2):
            for c, nf in fmts.items():
                row[c - 1].number_format = nf

    return len(src_rows), len(exploded)


# ---------- summary / PDF ----------

def fmt_num(x):
    if x == int(x):
        return f"{int(x)}"
    return f"{x:.3f}".rstrip("0").rstrip(".")


def _blank_agg():
    return {"sum": {}, "has": {}, "uom": None, "suom": None,
            "pkg_num": 0.0, "pkg_has": False, "pkg_numeric": True, "pkg_txt": set()}


def summarize(ws, cols):
    """
    Exploded sheet එක DISPLAY_ASN_NUMBER අනුව group කරනවා.
    Return: OrderedDict[asn] -> {"client": <code>, "groups": OrderedDict[(item,lot)] -> agg}
    """
    asns = OrderedDict()
    for r in range(2, ws.max_row + 1):
        if is_blank_row(ws, r):
            continue
        asn = ws.cell(row=r, column=cols["DISPLAY_ASN_NUMBER"]).value if cols.get("DISPLAY_ASN_NUMBER") else None
        client = ws.cell(row=r, column=cols["CLIENT_CODE"]).value if cols.get("CLIENT_CODE") else None
        item = ws.cell(row=r, column=cols["DISPLAY_ITEM_NUMBER"]).value if cols.get("DISPLAY_ITEM_NUMBER") else None
        lot = ws.cell(row=r, column=cols["LOT_NUMBER"]).value if cols.get("LOT_NUMBER") else None

        bucket = asns.setdefault(asn, {"client": client, "groups": OrderedDict()})
        if bucket["client"] in (None, "") and client not in (None, ""):
            bucket["client"] = client

        g = bucket["groups"].setdefault((item, lot), _blank_agg())
        for fld in ("QUANTITY", "S_QTY", "GROSS_WEIGHT", "NET_WEIGHT"):
            if cols.get(fld):
                x = to_num(ws.cell(row=r, column=cols[fld]).value)
                if x is not None:
                    g["sum"][fld] = g["sum"].get(fld, 0.0) + x
                    g["has"][fld] = True
        if cols.get("PACKAGE_TYPE"):
            pv = ws.cell(row=r, column=cols["PACKAGE_TYPE"]).value
            x = to_num(pv)
            if x is not None:
                g["pkg_num"] += x
                g["pkg_has"] = True
            elif pv not in (None, ""):
                g["pkg_numeric"] = False
                g["pkg_txt"].add(str(pv))
        if cols.get("UOM") and g["uom"] is None:
            g["uom"] = ws.cell(row=r, column=cols["UOM"]).value
        if cols.get("S_UOM") and g["suom"] is None:
            g["suom"] = ws.cell(row=r, column=cols["S_UOM"]).value
    return asns


def qr_flowable(data, size_mm=22):
    return Image(io.BytesIO(_qr_png(str(data) if data not in (None, "") else "N/A")),
                 width=size_mm * mm, height=size_mm * mm)


def _g_cell(g, fld):
    return fmt_num(g["sum"][fld]) if g["has"].get(fld) else "-"


def _pkg_cell(g):
    if g["pkg_numeric"]:
        return fmt_num(g["pkg_num"]) if g["pkg_has"] else "-"
    return ", ".join(sorted(g["pkg_txt"])) or "-"


def make_summary_pdf(ws, cols, doc_title="ASN Summary"):
    asns = summarize(ws, cols)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=14 * mm, bottomMargin=14 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm, title=doc_title)
    ss = getSampleStyleSheet()
    h = ParagraphStyle("h", parent=ss["Title"], fontSize=16, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=ss["Normal"], fontSize=9, textColor=colors.HexColor("#5a6b7b"), spaceAfter=8)
    lbl = ParagraphStyle("lbl", parent=ss["Normal"], fontSize=9, textColor=colors.grey)
    val = ParagraphStyle("val", parent=ss["Normal"], fontSize=13, leading=15)

    NAVY = colors.HexColor("#16324a")
    GRID = colors.HexColor("#b8c4cf")
    ZEBRA = colors.HexColor("#eef2f6")

    story = []

    # ---------- Page 1: Overview (සියලුම ASN) ----------
    story.append(Paragraph("ASN Overview", h))
    story.append(Paragraph(f"All DISPLAY_ASN_NUMBER &nbsp;·&nbsp; {len(asns)} ASN", sub))
    ov_header = ["DISPLAY_ASN_NUMBER", "DISPLAY_ITEM_NUMBER", "LOT_NUMBER", "QUANTITY\n(Total)", "UOM"]
    ov_rows = [ov_header]
    span_cmds = []
    ridx = 1
    for asn, bucket in asns.items():
        start = ridx
        for (item, lot), g in bucket["groups"].items():
            ov_rows.append([
                "" if asn is None else str(asn),
                "" if item is None else str(item),
                "" if lot in (None, "") else str(lot),
                _g_cell(g, "QUANTITY"),
                "" if g["uom"] is None else str(g["uom"]),
            ])
            ridx += 1
        if ridx - start > 1:  # එකම ASN එකේ rows කිහිපයක් → ASN cell merge
            span_cmds.append(("SPAN", (0, start), (0, ridx - 1)))
    ov = Table(ov_rows, repeatRows=1, colWidths=[42 * mm, 56 * mm, 28 * mm, 32 * mm, 28 * mm])
    ov.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTSIZE", (0, 0), (-1, 0), 8), ("FONTSIZE", (0, 1), (-1, -1), 8.5),
        ("GRID", (0, 0), (-1, -1), 0.4, GRID), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"), ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ZEBRA]),
        ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ] + span_cmds))
    story.append(ov)

    # ---------- ASN එකකට වෙන වෙනම page ----------
    det_header = ["DISPLAY_ITEM_NUMBER", "LOT_NUMBER", "QUANTITY\n(Total)", "UOM",
                  "S_QTY\n(Total)", "S_UOM", "PACKAGE_TYPE\n(Total)",
                  "GROSS_WEIGHT\n(Total)", "NET_WEIGHT\n(Total)"]
    for asn, bucket in asns.items():
        story.append(PageBreak())
        client = bucket["client"]
        story.append(Paragraph("ASN Summary", h))
        head = Table([[
            [Paragraph("CLIENT_CODE", lbl),
             Paragraph(f"<b>{'' if client is None else client}</b>", val), Spacer(1, 3), qr_flowable(client)],
            [Paragraph("DISPLAY_ASN_NUMBER", lbl),
             Paragraph(f"<b>{'' if asn is None else asn}</b>", val), Spacer(1, 3), qr_flowable(asn)],
        ]], colWidths=[93 * mm, 93 * mm])
        head.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("LEFTPADDING", (0, 0), (-1, -1), 8), ("RIGHTPADDING", (0, 0), (-1, -1), 8),
            ("TOPPADDING", (0, 0), (-1, -1), 8), ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        story += [head, Spacer(1, 10)]

        rows = [det_header]
        for (item, lot), g in bucket["groups"].items():
            rows.append([
                "" if item is None else str(item),
                "" if lot in (None, "") else str(lot),
                _g_cell(g, "QUANTITY"), "" if g["uom"] is None else str(g["uom"]),
                _g_cell(g, "S_QTY"), "" if g["suom"] is None else str(g["suom"]),
                _pkg_cell(g), _g_cell(g, "GROSS_WEIGHT"), _g_cell(g, "NET_WEIGHT"),
            ])
        t = Table(rows, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), NAVY), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, 0), 7.5), ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.4, GRID),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ZEBRA]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (2, 0), (-1, -1), "RIGHT"), ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(t)

    doc.build(story)
    buf.seek(0)
    n_groups = sum(len(b["groups"]) for b in asns.values())
    return buf.getvalue(), len(asns), n_groups


# ---------- HU_ID labels (barcode + QR) ----------

LABEL_DETAIL_FIELDS = ["DISPLAY_ITEM_NUMBER", "DISPLAY_ASN_NUMBER", "LOT_NUMBER",
                       "QUANTITY", "SUPPLIER_DESC"]


@lru_cache(maxsize=8192)
def _qr_png(data):
    qr = qrcode.QRCode(border=1, box_size=10)
    qr.add_data(data)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    b = io.BytesIO()
    img.save(b, format="PNG")
    return b.getvalue()


def _qr_image(data, size_mm=18):
    return Image(io.BytesIO(_qr_png(str(data) if data not in (None, "") else "N/A")),
                 width=size_mm * mm, height=size_mm * mm)


def _barcode(value, width_mm=58, h_mm=11):
    return createBarcodeDrawing("Code128", value=str(value) if value not in (None, "") else "N/A",
                                barHeight=h_mm * mm, width=width_mm * mm, humanReadable=True, fontSize=6)


def make_labels_pdf(records, per_row=2):
    """records: [{'hu':..., 'details':[(label,val),...]}, ...]"""
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=8 * mm, bottomMargin=8 * mm,
                            leftMargin=8 * mm, rightMargin=8 * mm, title="HU_ID Labels")
    ss = getSampleStyleSheet()
    big = ParagraphStyle("big", parent=ss["Normal"], fontSize=10, leading=12)
    det = ParagraphStyle("det", parent=ss["Normal"], fontSize=7.5, leading=9)

    col_w = (190.0 / per_row)
    qr_mm = 16 if per_row >= 3 else 18
    bc_mm = col_w - 24

    def cell(rec):
        left = [Paragraph(f"<b>{rec['hu']}</b>", big)]
        for k, v in rec["details"]:
            if v not in (None, ""):
                left.append(Paragraph(f"<font color='#667'>{k}:</font> {v}", det))
        top = Table([[left, _qr_image(rec["hu"], qr_mm)]], colWidths=[(col_w - qr_mm - 6) * mm, qr_mm * mm])
        top.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP"),
                                 ("LEFTPADDING", (0, 0), (-1, -1), 2), ("RIGHTPADDING", (0, 0), (-1, -1), 2),
                                 ("TOPPADDING", (0, 0), (-1, -1), 1), ("BOTTOMPADDING", (0, 0), (-1, -1), 1)]))
        inner = Table([[top], [_barcode(rec["hu"], bc_mm)]], colWidths=[col_w * mm])
        inner.setStyle(TableStyle([("ALIGN", (0, 1), (0, 1), "CENTER"),
                                   ("TOPPADDING", (0, 0), (-1, -1), 1), ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
                                   ("LEFTPADDING", (0, 0), (-1, -1), 3)]))
        return inner

    cells = [cell(r) for r in records]
    rows = []
    for i in range(0, len(cells), per_row):
        row = cells[i:i + per_row]
        while len(row) < per_row:
            row.append("")
        rows.append(row)
    grid = Table(rows, colWidths=[col_w * mm] * per_row)
    grid.setStyle(TableStyle([("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#aab")),
                              ("VALIGN", (0, 0), (-1, -1), "TOP"),
                              ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 6)]))
    doc.build([grid])
    buf.seek(0)
    return buf.getvalue()


def build_label_records(ws, hu_col, detail_cols, limit):
    """Exploded sheet එකෙන් label records හදනවා (HU_ID + selected details)."""
    records = []
    truncated = False
    for r in range(2, ws.max_row + 1):
        if is_blank_row(ws, r):
            continue
        if len(records) >= limit:
            truncated = True
            break
        hu = ws.cell(row=r, column=hu_col).value if hu_col else None
        details = []
        for fld, ci in detail_cols:
            if ci:
                details.append((fld, ws.cell(row=r, column=ci).value))
        records.append({"hu": "" if hu is None else str(hu), "details": details})
    return records, truncated


# ---------- UI ----------

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&family=JetBrains+Mono:wght@500;700&display=swap');

:root{ --navy:#16324a; --navy-d:#0e2438; --amber:#f5a623; --ink:#1a2733; --muted:#5a6b7b; --line:#e3eaf0; }

html, body, [class*="css"]{ font-family:'Inter',system-ui,sans-serif; }
.stApp{ background:linear-gradient(180deg,#f7fafc 0%,#eef3f7 100%); }
.block-container{ padding-top:1.4rem; animation:fadeUp .5s ease both; }

@keyframes fadeUp{ from{opacity:0; transform:translateY(10px);} to{opacity:1; transform:translateY(0);} }
@keyframes sheen{ 0%{background-position:0% 50%;} 100%{background-position:200% 50%;} }

/* hero */
.hero{ position:relative; overflow:hidden; border-radius:16px; padding:26px 28px;
  background:linear-gradient(110deg,#0e2438,#16324a 60%,#1b3f5c 100%);
  box-shadow:0 10px 30px -12px rgba(14,36,56,.55); color:#fff; }
.hero::after{ content:""; position:absolute; top:0; left:0; right:0; height:3px;
  background:linear-gradient(90deg,transparent,var(--amber),transparent);
  background-size:200% 100%; animation:sheen 3.5s linear infinite; }
.hero .eyebrow{ font-family:'JetBrains Mono',monospace; font-size:.72rem; letter-spacing:.22em;
  text-transform:uppercase; color:var(--amber); margin:0 0 6px; }
.hero h1{ font-size:1.8rem; font-weight:700; margin:0; line-height:1.1; }
.hero p{ margin:.5rem 0 0; color:#bcd0e2; font-size:.92rem; max-width:60ch; }
.hero .pills{ margin-top:14px; display:flex; gap:8px; flex-wrap:wrap; }
.hero .pill{ font-family:'JetBrains Mono',monospace; font-size:.7rem; letter-spacing:.04em;
  background:rgba(255,255,255,.08); border:1px solid rgba(255,255,255,.16);
  padding:5px 10px; border-radius:999px; color:#dbe7f2; }

/* section headers */
h2, h3{ color:var(--navy)!important; font-weight:600!important; }

/* metrics as cards */
[data-testid="stMetric"]{ background:#fff; border:1px solid var(--line); border-left:4px solid var(--amber);
  border-radius:12px; padding:14px 16px; box-shadow:0 4px 14px -10px rgba(22,50,74,.4);
  transition:transform .15s ease, box-shadow .15s ease; animation:fadeUp .5s ease both; }
[data-testid="stMetric"]:hover{ transform:translateY(-2px); box-shadow:0 10px 22px -12px rgba(22,50,74,.5); }
[data-testid="stMetricValue"]{ font-family:'JetBrains Mono',monospace; color:var(--navy); }

/* buttons */
.stButton>button, [data-testid="stDownloadButton"]>button{
  border-radius:10px; font-weight:600; border:1px solid var(--line);
  transition:transform .12s ease, box-shadow .18s ease, background .2s ease; }
.stButton>button:hover, [data-testid="stDownloadButton"]>button:hover{
  transform:translateY(-2px); box-shadow:0 8px 18px -10px rgba(22,50,74,.55); }
button[kind="primary"]{ background:linear-gradient(120deg,#16324a,#205074)!important; border:none!important; }
button[kind="primary"]:hover{ background:linear-gradient(120deg,#1b3f5c,#27618d)!important; }
[data-testid="stDownloadButton"]>button{ background:#fff; color:var(--navy); }

/* file uploader */
[data-testid="stFileUploader"]{ background:#fff; border:1.5px dashed #c4d3e0; border-radius:14px;
  padding:8px 12px; transition:border-color .2s ease, box-shadow .2s ease; }
[data-testid="stFileUploader"]:hover{ border-color:var(--amber); box-shadow:0 6px 18px -12px rgba(245,166,35,.5); }

/* code-ish captions */
.stCaption, [data-testid="stCaptionContainer"]{ color:var(--muted)!important; }

/* widget labels + radio/checkbox option text — read වෙන්න dark කරනවා */
[data-testid="stWidgetLabel"], [data-testid="stWidgetLabel"] *,
.stRadio label, .stRadio label *,
.stCheckbox label, .stCheckbox label *,
.stSelectbox label, .stMultiSelect label,
.stTextInput label, .stNumberInput label,
[data-baseweb="radio"] *, [data-testid="stRadio"] label, [data-testid="stRadio"] label *,
[data-testid="stCheckbox"] label, [data-testid="stCheckbox"] label *{
  color:var(--ink)!important;
}
/* selected radio / checkbox tick → navy (off-brand red නැති කරනවා) */
[data-baseweb="radio"] [aria-checked="true"]>div:first-child{
  background-color:var(--navy)!important; border-color:var(--navy)!important; }
[data-baseweb="checkbox"] [aria-checked="true"]{
  background-color:var(--navy)!important; border-color:var(--navy)!important; }

@media (prefers-reduced-motion: reduce){
  *{ animation:none!important; transition:none!important; }
}
</style>
""", unsafe_allow_html=True)

st.markdown("""
<div class="hero">
  <p class="eyebrow">Warehouse · ASN Toolkit</p>
  <h1>ASN S_QTY Exploder</h1>
  <p>S_QTY අගයෙන් line ගාන හදනවා — S_QTY 10 → line 10ක්, හැම line එකකම S_QTY = 1.
     Exploded Excel · ASN Summary PDF · HU_ID barcode + QR labels.</p>
  <div class="pills">
    <span class="pill">EXPLODE</span>
    <span class="pill">HU_ID GEN</span>
    <span class="pill">SUMMARY PDF</span>
    <span class="pill">BARCODE + QR</span>
  </div>
</div>
""", unsafe_allow_html=True)
st.write("")

uploaded = st.file_uploader("Excel file එකක් upload කරන්න (.xlsx)", type=["xlsx"])
if uploaded is None:
    st.info("⬆️ GENERAL Format හරි ATTRIBUTE EXTENTD format හරි Excel එකක් upload කරන්න.")
    st.stop()

raw = uploaded.read()
wb = load_workbook(io.BytesIO(raw))

sheet_info = {}
for name in wb.sheetnames:
    ws = wb[name]
    sheet_info[name] = {
        "sqty": find_col(ws, "S_QTY"),
        "line": find_col(ws, "ASN_LINE_NUMBER"),
        "hu": find_col(ws, "HU_ID"),
        "item": find_col(ws, "DISPLAY_ITEM_NUMBER"),
        "rows": max(ws.max_row - 1, 0),
    }

sqty_sheets = [n for n, i in sheet_info.items() if i["sqty"]]
if not sqty_sheets:
    st.error("මේ file එකේ කිසිම sheet එකක `S_QTY` column එකක් නෑ. Header row එක row 1 ද කියලා බලන්න.")
    st.stop()

col1, col2 = st.columns([2, 1])
with col1:
    target_sheet = st.selectbox("Explode කරන්න ඕන sheet එක", options=sqty_sheets, index=0)
with col2:
    renumber = st.checkbox("ASN_LINE_NUMBER අලුතෙන් 1..N", value=True)

info = sheet_info[target_sheet]
ws_target = wb[target_sheet]

if len(sqty_sheets) == 1:
    st.caption(f"✓ `S_QTY` column එක හම්බුණේ sheet **{target_sheet}** එකේ (sheet name එක වෙනස් වුණත් auto-detect වෙනවා).")
else:
    st.caption(f"ℹ️ `S_QTY` column එක තියෙන sheets {len(sqty_sheets)}ක්: {', '.join(sqty_sheets)} — එකක් තෝරන්න.")

# ---- HU_ID ----
st.subheader("🏷️ HU_ID generate")
if info["hu"] is None:
    st.warning("මේ sheet එකේ `HU_ID` column එකක් නෑ — HU_ID generate skip වෙනවා.")
    hu_cfg = {"mode": "keep"}
else:
    hu_mode_label = st.radio(
        "HU_ID හදන විදිය",
        options=["Keep original (නොවෙනස්ව)", "None (හිස්ව)",
                 "Letters + Number (1,2,3..)", "DISPLAY_ITEM_NUMBER + Number (1,2,3..)"],
        index=0,
    )
    mode_map = {
        "Keep original (නොවෙනස්ව)": "keep",
        "None (හිස්ව)": "none",
        "Letters + Number (1,2,3..)": "letters",
        "DISPLAY_ITEM_NUMBER + Number (1,2,3..)": "item",
    }
    mode = mode_map[hu_mode_label]
    hu_cfg = {"mode": mode, "letters": "", "number": "", "psep": "", "sep": "", "start": 1, "pad": 0}
    if mode in ("letters", "item"):
        if mode == "letters":
            r1c1, r1c2, r1c3 = st.columns(3)
            with r1c1:
                hu_cfg["letters"] = st.text_input("Letters (prefix)", value="HKEFL")
            with r1c2:
                hu_cfg["psep"] = st.text_input("Prefix–Number අතර", value="",
                                               help="Letters එකයි Number එකයි අතර (උදා: - ). හිස් නම් join වෙනවා.")
            with r1c3:
                hu_cfg["number"] = st.text_input("Number (fixed)", value="260613")
        else:
            r1c1, r1c2, r1c3 = st.columns(3)
            with r1c1:
                st.text_input("Prefix", value="DISPLAY_ITEM_NUMBER", disabled=True,
                              help="හැම row එකකම DISPLAY_ITEM_NUMBER value එක prefix විදිහට යනවා")
            with r1c2:
                hu_cfg["psep"] = st.text_input("Prefix–Number අතර", value="-",
                                               help="Item එකයි Number එකයි අතර (උදා: - ).")
            with r1c3:
                hu_cfg["number"] = st.text_input("Number (fixed)", value="260613")

        r2c1, r2c2, r2c3 = st.columns(3)
        with r2c1:
            hu_cfg["sep"] = st.text_input("Number–Line අතර", value=".",
                                          help="Number එකයි line අංකයයි අතර (උදා: . )")
        with r2c2:
            hu_cfg["start"] = st.number_input("Line අංකය පටන් (1,2,3..)", value=1, step=1)
        with r2c3:
            hu_cfg["pad"] = st.number_input("Zero-pad", value=0, min_value=0, step=1,
                                            help="line අංකයට. උදා: 2 නම් 1 → 01. 0 නම් pad නෑ.")

        item_example = None
        if mode == "item" and info["item"]:
            item_example = ws_target.cell(row=2, column=info["item"]).value
        prev = []
        for k in range(3):
            pfx = hu_cfg["letters"] if mode == "letters" else ("" if item_example is None else str(item_example))
            prev.append(build_hu(pfx, hu_cfg["psep"], hu_cfg["number"], hu_cfg["sep"],
                                 int(hu_cfg["start"]) + k, hu_cfg["pad"]))
        st.caption("Preview: " + "  ·  ".join(f"`{p}`" for p in prev) + "  · …")

# ---- Summary PDF ----
st.subheader("📄 Summary PDF")
make_pdf = st.checkbox("Summary PDF එකකුත් හදන්න", value=True)
sum_cols = {f: find_col(ws_target, f) for f in SUMMARY_FIELDS}
if make_pdf:
    missing = [f for f in SUMMARY_FIELDS if not sum_cols[f]]
    if missing:
        st.caption("⚠️ නැති columns (skip වෙනවා): " + ", ".join(missing))
    else:
        st.caption("CLIENT_CODE + QR · DISPLAY_ASN_NUMBER + QR · item-wise totals (QUANTITY, S_QTY, weights…)")

# ---- HU_ID Labels (barcode + QR) — generate වෙන්නේ results වලට පස්සේ, ඕනේ නම් විතරක් ----
st.subheader("🏷️ HU_ID Labels (barcode + QR)")
if info["hu"] is None:
    st.caption("මේ sheet එකේ `HU_ID` column එකක් නෑ — labels හදන්න බෑ.")
else:
    st.caption("Explode කළාට පස්සේ පහළින් **ඕනේ නම් විතරක්** barcode + QR labels PDF එක හදාගන්න පුළුවන්.")

# ---- metrics ----
preview_total = 0
for r in range(2, ws_target.max_row + 1):
    if is_blank_row(ws_target, r):
        continue
    preview_total += to_int_qty(ws_target.cell(row=r, column=info["sqty"]).value)

m1, m2, m3 = st.columns(3)
m1.metric("දැනට data rows", info["rows"])
m2.metric("Explode වුණාම rows", preview_total)
m3.metric("අලුතෙන් එකතු වන rows", preview_total - info["rows"])

st.divider()

if st.button("🚀 Explode & Generate", type="primary", use_container_width=True):
    n_src, n_out = explode_sheet(
        ws_target, sqty_col=info["sqty"], line_col=info["line"], renumber_line=renumber,
        hu_col=info["hu"], item_col=info["item"], hu_cfg=hu_cfg,
    )
    out_buf = io.BytesIO()
    wb.save(out_buf)

    base = uploaded.name.rsplit(".", 1)[0]

    pdf_bytes, n_asn, ngrp, pdf_err = None, 0, 0, None
    if make_pdf:
        try:
            pdf_bytes, n_asn, ngrp = make_summary_pdf(ws_target, sum_cols)
        except Exception as e:
            pdf_err = str(e)

    # Labels: PDF මෙතනදී හදන්නේ නෑ (speed) — records විතරක් capture කරනවා, generate වෙන්නේ ඕනේ නම් පහළින්
    hu_records = None
    if info["hu"]:
        try:
            detail_cols = [(f, find_col(ws_target, f)) for f in LABEL_DETAIL_FIELDS]
            hu_records, _ = build_label_records(ws_target, info["hu"], detail_cols, 20000)
        except Exception:
            hu_records = None
    st.session_state.pop("labels_pdf", None)  # පරණ labels clear

    headers = [ws_target.cell(row=1, column=c).value for c in range(1, ws_target.max_column + 1)]
    preview = []
    for r in range(2, min(ws_target.max_row, 16) + 1):
        preview.append([ws_target.cell(row=r, column=c).value for c in range(1, ws_target.max_column + 1)])

    # download buttons click එකකින් rerun වුණත් නැති නොවෙන්න session_state එකේ තියාගන්නවා
    st.session_state["result"] = {
        "base": base, "n_src": n_src, "n_out": n_out, "sheet": target_sheet,
        "excel": out_buf.getvalue(), "pdf": pdf_bytes, "ngrp": ngrp, "n_asn": n_asn, "pdf_err": pdf_err,
        "hu_records": hu_records,
        "headers": headers, "preview": preview,
    }

# ---- results (session_state එකෙන් — rerun වුණත් download buttons නැති වෙන්නේ නෑ) ----
res = st.session_state.get("result")
if res:
    st.success(f"✅ සාර්ථකයි · {res['n_src']} rows → {res['n_out']} rows ({res['sheet']})")

    d1, d2 = st.columns(2)
    with d1:
        st.download_button("⬇️ Exploded Excel", data=res["excel"],
                           file_name=f"{res['base']}_EXPLODED.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True, key="dl_excel")
    with d2:
        if res["pdf"] is not None:
            st.download_button(f"⬇️ Summary PDF ({res.get('n_asn', 0)} ASN)", data=res["pdf"],
                               file_name=f"{res['base']}_SUMMARY.pdf", mime="application/pdf",
                               use_container_width=True, key="dl_pdf")
        elif res["pdf_err"]:
            st.error(f"PDF error: {res['pdf_err']}")

    # ---- HU_ID Labels — ඕනේ නම් විතරක් generate + download ----
    recs_all = res.get("hu_records")
    if recs_all:
        st.divider()
        st.markdown("### 🏷️ HU_ID Labels (barcode + QR)")
        st.caption("ඕනේ නම් විතරක් — options තෝරලා **Labels PDF හදන්න** click කරන්න.")
        total = len(recs_all)
        lc1, lc2, lc3 = st.columns([1, 1, 1.3])
        with lc1:
            per_row = st.selectbox("පේළියකට labels", options=[2, 3], index=0, key="lbl_per_row")
        with lc2:
            limit = st.number_input("උපරිම labels (speed)", min_value=1, value=min(500, total),
                                    step=100, key="lbl_limit",
                                    help="ලොකු ගානකට (>1000) තත්පර කීපයක් යනවා.")
        with lc3:
            st.write("")
            if st.button("🏷️ Labels PDF හදන්න", use_container_width=True, key="gen_labels"):
                recs = recs_all[:int(limit)]
                with st.spinner(f"Labels {len(recs)}ක් හදනවා…"):
                    st.session_state["labels_pdf"] = make_labels_pdf(recs, per_row=per_row)
                    st.session_state["labels_n"] = len(recs)

        if st.session_state.get("labels_pdf"):
            ln = st.session_state.get("labels_n", 0)
            st.download_button(f"⬇️ HU_ID Labels PDF ({ln})", data=st.session_state["labels_pdf"],
                               file_name=f"{res['base']}_LABELS.pdf", mime="application/pdf",
                               use_container_width=True, key="dl_labels")
            if ln < total:
                st.caption(f"ℹ️ {total}න් {ln}ක් විතරයි හැදුවේ. ඔක්කොම ඕන නම් 'උපරිම labels' වැඩි කරලා ආයෙ හදන්න.")

    st.divider()
    st.caption("Output preview (මුල් rows 15)")
    st.dataframe({(hh or f"col{i}"): [row[i] for row in res["preview"]]
                  for i, hh in enumerate(res["headers"])},
                 use_container_width=True, height=360)
